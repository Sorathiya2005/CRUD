from django.shortcuts import redirect,render
from django.views import View
from .models import Student
from .forms import AddStudentForm
import openpyxl


# Create your views here.
class Home(View):
    def get(self,request):
        stu_data=Student.objects.all()
        return render(request,'core/home.html',{'studata':stu_data})

class Add_Student(View):
    def get(self,request):
        fm = AddStudentForm()
        return render(request,'core/add-student.html',{'form':fm})

    def post(self,request):
        fm = AddStudentForm(request.POST)
        if fm.is_valid():
            fm.save()
            return redirect('/')
        else:
            return render(request,'core/add-student.html',{'form':fm})

class Delete_Student(View):
    def post(self,request):
        data=request.POST
        id=data.get('id')
        studata = Student.objects.get(id=id)
        studata.delete()
        return redirect ('/')

class Add_Excel(View):
    pass


def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data":excel_data})
